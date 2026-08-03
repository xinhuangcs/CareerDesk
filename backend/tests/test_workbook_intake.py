
import asyncio
from datetime import time
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

from careerdesk.core.config import get_settings
from careerdesk.agentic.tools.manage_jobs import ParseJobsTool
from careerdesk.features.applications.workbook_intake import (
    STANDARD_HEADERS,
    parse_standard_workbook,
    read_workbook,
    standard_positions_from_structured_text,
)
from careerdesk.features.applications.intake_models import MAX_BATCH_POSITIONS


def test_standard_csv_fails_closed_per_row_and_preserves_metadata(tmp_path: Path):
    source = tmp_path / "jobs.csv"
    source.write_text(
        "公司,职位,状态,申请日期,下一步,待办日期,待办时间,备注,星标\n"
        "星海科技,数据分析师,面试,2026/07/02,终面,2026-07-25,09:30,准备业务案例,是\n"
        ",缺公司,已投递,2026-07-03,,,,,否\n",
        encoding="utf-8",
    )

    parsed = parse_standard_workbook(source)

    assert parsed.is_standard is True
    assert parsed.source_rows == 2
    assert parsed.skipped_rows == 1
    assert len(parsed.positions) == 1
    position = parsed.positions[0]
    assert position.company == "星海科技"
    assert position.position == "数据分析师"
    assert position.stage == "interviewing"
    assert position.applied_date == "2026-07-02"
    assert position.next_action is not None
    assert position.next_action.time == "09:30"
    assert position.application_note == "准备业务案例"
    assert position.priority == "high"
    assert position.source_row == "jobs!2"

    canonical = standard_positions_from_structured_text(parsed.structured_text)
    assert canonical is not None
    assert canonical.source_rows == 2
    assert canonical.skipped_rows == 1
    assert canonical.positions == parsed.positions


def test_annotated_template_headers_map_without_guessing(tmp_path: Path):
    source = tmp_path / "annotated.csv"
    source.write_text(
        "投递日期（可选）,公司名称,岗位名称,部门（可选）,投递渠道（可选）,"
        "优先级（可选）,当前阶段（可选）,当前环节（可选）,下一阶段（可选）,"
        "下一环节（可选）,下一环节的日期（可选）,下一环节的时间（可选）\n"
        "2026-07-02,星海科技,数据分析师,数据组,内推,高,面试中,二面,Offer,终面,"
        "2026-07-25,14:00\n",
        encoding="utf-8",
    )

    parsed = parse_standard_workbook(source)

    assert len(parsed.positions) == 1
    assert parsed.positions[0].applied_date == "2026-07-02"
    assert parsed.positions[0].channel == "内推"
    assert parsed.positions[0].current_step == "二面"
    assert parsed.positions[0].priority == "high"
    assert parsed.positions[0].next_action is not None
    assert parsed.positions[0].next_action.stage == "offer"
    assert parsed.positions[0].next_action.step == "终面"
    assert parsed.positions[0].next_action.date == "2026-07-25"
    assert parsed.positions[0].next_action.time == "14:00"


def test_unknown_explicit_next_stage_drops_next_action_instead_of_guessing(tmp_path: Path):
    source = tmp_path / "unknown-next-stage.csv"
    source.write_text(
        "公司名称,岗位名称,当前阶段,下一环节,下一阶段\n"
        "星海科技,数据分析师,面试中,终面,大概下一轮\n",
        encoding="utf-8",
    )

    parsed = parse_standard_workbook(source)

    assert len(parsed.positions) == 1
    assert parsed.positions[0].stage == "interviewing"
    assert parsed.positions[0].next_action is None


def test_missing_required_fields_skip_only_that_row(tmp_path: Path):
    source = tmp_path / "partial.csv"
    source.write_text(
        "公司名称,岗位名称,渠道\n"
        "甲公司,前端工程师,官网\n"
        ",缺少公司,内推\n"
        "缺少岗位公司,,猎头\n"
        "乙公司,后端工程师,招聘会\n",
        encoding="utf-8",
    )

    parsed = parse_standard_workbook(source)

    assert parsed.is_standard is True
    assert parsed.source_rows == 4
    assert parsed.skipped_rows == 2
    assert [
        (position.company, position.position, position.channel, position.source_row)
        for position in parsed.positions
    ] == [
        ("甲公司", "前端工程师", "官网", "partial!2"),
        ("乙公司", "后端工程师", "招聘会", "partial!5"),
    ]


def test_xlsx_keeps_each_sheet_headers_and_nonstandard_rows_for_agent(tmp_path: Path):
    source = tmp_path / "mixed.xlsx"
    workbook = Workbook()
    notes = workbook.active
    notes.title = "说明"
    notes.append(["使用说明"])
    notes.append(["这一页不是岗位"])
    jobs = workbook.create_sheet("岗位")
    jobs.append(["company", "job title", "stage"])
    jobs.append(["远景网络", "后端工程师", "applied"])
    workbook.save(source)

    parsed = parse_standard_workbook(source)

    assert parsed.is_standard is True
    assert parsed.source_rows == 2
    assert parsed.skipped_rows == 1
    assert [position.company for position in parsed.positions] == ["远景网络"]
    assert "[工作表 说明，第 2 行]" in parsed.structured_text
    assert "[工作表 岗位，第 2 行]" in parsed.structured_text


def test_xlsx_preserves_physical_row_and_native_excel_time(tmp_path: Path):
    source = tmp_path / "native-values.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "岗位"
    sheet.append(["公司名称", "岗位名称", "下一步", "下一步日期", "下一步时间"])
    sheet.append([None, None, None, None, None])
    sheet.append(["云杉科技", "平台工程师", "技术面", "2026-08-01", time(9, 45)])
    workbook.save(source)

    parsed = parse_standard_workbook(source)

    assert len(parsed.positions) == 1
    assert parsed.positions[0].source_row == "岗位!3"
    assert parsed.positions[0].next_action is not None
    assert parsed.positions[0].next_action.time == "09:45"
    assert "[工作表 岗位，第 3 行]" in parsed.structured_text


def test_duplicate_semantic_headers_are_not_silently_selected(tmp_path: Path):
    source = tmp_path / "ambiguous.csv"
    source.write_text(
        "公司名称,公司,岗位名称\n甲公司,乙公司,工程师\n",
        encoding="utf-8",
    )

    parsed = parse_standard_workbook(source)

    assert parsed.is_standard is False
    assert parsed.positions == ()
    assert parsed.skipped_rows == 1
    assert standard_positions_from_structured_text(parsed.structured_text) is None


def test_batch_limit_is_200_and_excess_rows_are_reported(tmp_path: Path):
    source = tmp_path / "capacity.csv"
    rows = ["公司名称,岗位名称"] + [
        f"测试公司{index:03d},测试岗位{index:03d}"
        for index in range(1, MAX_BATCH_POSITIONS + 2)
    ]
    source.write_text("\n".join(rows) + "\n", encoding="utf-8")

    parsed = parse_standard_workbook(source)

    assert MAX_BATCH_POSITIONS == 200
    assert len(parsed.positions) == 200
    assert parsed.source_rows == 201
    assert parsed.skipped_rows == 1
    assert parsed.structured_text.startswith("[CAREERDESK_STANDARD_ROWS_V1]\n")
    canonical = standard_positions_from_structured_text(parsed.structured_text)
    assert canonical is not None
    assert len(canonical.positions) == 200
    assert canonical.source_rows == 201
    assert canonical.skipped_rows == 1


def test_nonstandard_rows_do_not_consume_import_capacity(tmp_path: Path):
    source = tmp_path / "instructions-before-jobs.xlsx"
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "说明"
    instructions.append(["说明"])
    for index in range(MAX_BATCH_POSITIONS):
        instructions.append([f"第 {index + 1} 行说明，不是岗位"])
    jobs = workbook.create_sheet("岗位")
    jobs.append(["公司名称", "岗位名称"])
    jobs.append(["远景网络", "后端工程师"])
    workbook.save(source)

    parsed = parse_standard_workbook(source)

    assert parsed.is_standard is True
    assert parsed.source_rows == MAX_BATCH_POSITIONS + 1
    assert parsed.skipped_rows == MAX_BATCH_POSITIONS
    assert [(item.company, item.position) for item in parsed.positions] == [
        ("远景网络", "后端工程师"),
    ]


def test_checked_in_example_uses_the_declared_user_friendly_headers():
    source = Path(__file__).parents[2] / "frontend/public/careerdesk-job-import-example-zh-CN.xlsx"

    workbook = read_workbook(source)
    parsed = parse_standard_workbook(source)

    assert workbook.headers[:len(STANDARD_HEADERS)] == STANDARD_HEADERS
    assert not any(workbook.headers[len(STANDARD_HEADERS):])
    assert STANDARD_HEADERS[0] == "投递日期（可选）"
    assert STANDARD_HEADERS[1:3] == ("公司名称", "岗位名称")
    assert "泡池原因（仅泡池子，可选）" not in STANDARD_HEADERS
    assert STANDARD_HEADERS[4] == "投递渠道（可选）"
    assert STANDARD_HEADERS[7:10] == (
        "优先级（可选）", "当前阶段（可选）", "当前环节（可选）",
    )
    assert STANDARD_HEADERS[10:14] == (
        "下一阶段（可选）", "下一环节（可选）",
        "下一环节的日期（可选）", "下一环节的时间（可选）",
    )
    assert len(parsed.positions) == 4
    assert parsed.skipped_rows == 0


def test_checked_in_english_example_round_trips_through_aliases():
    source = Path(__file__).parents[2] / "frontend/public/careerdesk-job-import-example-en.xlsx"

    workbook = read_workbook(source)
    parsed = parse_standard_workbook(source)

    assert workbook.headers[:3] == (
        "Application Date (optional)", "Company", "Role Title",
    )
    assert parsed.is_standard is True
    assert len(parsed.positions) == 4
    assert parsed.skipped_rows == 0


def test_timeline_file_import_needs_no_llm_and_applies_all_supported_fields(tmp_path, monkeypatch):
    monkeypatch.setenv("APP_DATA_DIR", str(tmp_path / "data"))
    monkeypatch.setenv("APP_LLM_MODEL", "")
    get_settings.cache_clear()
    from careerdesk.bootstrap.app import create_app

    csv_body = (
        "公司名称,岗位名称,部门,渠道,当前阶段,当前环节,投递日期,完成后阶段,下一步,"
        "下一步日期,下一步时间,下一步说明,岗位描述,岗位备注,重点岗位\n"
        "示例科技,产品经理,增长产品,内推,面试中,二面,2026-07-01,面试中,终面,"
        "2026-07-25,14:00,线上沟通,负责增长产品,准备增长案例,是\n"
        ",缺少公司,研发,官网,已投递,,,,,,,,,,否\n"
    ).encode()
    try:
        with TestClient(create_app()) as client:
            prepared = client.post(
                "/api/timeline/intake-operations/file",
                files={"file": ("jobs.csv", csv_body, "text/csv")},
            )
            assert prepared.status_code == 200
            preview = prepared.json()
            assert preview["status"] == "preview"
            assert preview["source_rows"] == 2
            assert preview["skipped_rows"] == 1
            assert len(preview["positions"]) == 1

            approved = client.post(
                f"/api/timeline/intake-operations/{preview['operation_id']}/approve",
                json={"exclude_indexes": []},
            )
            assert approved.status_code == 200
            assert approved.json()["state"] == "completed"

            board = client.get("/api/timeline/board").json()
            item = next(value for values in board["columns"].values() for value in values)
            detail = client.get(f"/api/timeline/applications/{item['id']}").json()
            assert detail["stage"] == "interviewing"
            assert detail["current_step"] == "二面"
            assert detail["application_note"] == "准备增长案例"
            assert detail["priority"] == "high"
            assert detail["next_action"] == {
                "stage": "interviewing", "step": "终面", "date": "2026-07-25",
                "time": "14:00", "note": "线上沟通",
            }
            assert detail["jd_text"] == "负责增长产品"
            assert detail["timeline_entries"][0]["occurred_date"] is None
    finally:
        get_settings.cache_clear()


def test_nonstandard_file_is_not_guessed_by_timeline_import(tmp_path, monkeypatch):
    monkeypatch.setenv("APP_DATA_DIR", str(tmp_path / "data"))
    monkeypatch.setenv("APP_LLM_MODEL", "")
    get_settings.cache_clear()
    from careerdesk.bootstrap.app import create_app

    try:
        with TestClient(create_app()) as client:
            response = client.post(
                "/api/timeline/intake-operations/file",
                files={"file": ("notes.csv", "求职记录\n昨天投了某个岗位\n".encode(), "text/csv")},
            )
            assert response.status_code == 200
            assert response.json()["status"] == "unrecognized"
            assert client.get("/api/timeline/intake-operations/pending").json() == {
                "operations": [],
            }
    finally:
        get_settings.cache_clear()


def test_batch_text_budget_skips_only_the_record_that_would_overflow(tmp_path: Path):
    source = tmp_path / "large-jd.csv"
    oversized_batch = (
        "公司名称,岗位名称,岗位描述\n"
        f"甲公司,前端工程师,{'甲' * 30_000}\n"
        f"乙公司,后端工程师,{'乙' * 30_001}\n"
        "丙公司,产品经理,短 JD\n"
    )
    source.write_text(oversized_batch, encoding="utf-8")

    parsed = parse_standard_workbook(source)

    assert [(item.company, item.position) for item in parsed.positions] == [
        ("甲公司", "前端工程师"),
        ("丙公司", "产品经理"),
    ]
    assert parsed.source_rows == 3
    assert parsed.skipped_rows == 1


def test_agent_standard_attachment_bypasses_llm_parser(tmp_path: Path):
    source = tmp_path / "jobs.csv"
    source.write_text("公司名称,岗位名称\n星海科技,数据分析师\n", encoding="utf-8")
    structured = parse_standard_workbook(source).structured_text
    calls = []

    class Service:
        def parse_standard_positions(self, user_id, positions, **metadata):
            calls.append((user_id, positions, metadata))
            return {
                "status": "preview",
                "positions": [{
                    "company": positions[0].company,
                    "position": positions[0].position,
                    "stage": "backlog",
                    "department": None,
                    "channel": None,
                    "next_action": None,
                    "already_exists": False,
                }],
            }

        async def parse_batch(self, *_args, **_kwargs):
            raise AssertionError("标准表格不得调用 LLM 解析器")

    response = asyncio.run(ParseJobsTool(Service(), "u1").arun({"text": structured}))

    assert response.status == "success"
    assert len(calls) == 1
    assert calls[0][2]["source_rows"] == 1
    assert calls[0][2]["skipped_rows"] == 0


def test_agent_nonstandard_attachment_falls_back_to_llm_parser(tmp_path: Path):
    source = tmp_path / "personal-tracker.csv"
    source.write_text(
        "我的记录,目标\n"
        "上周通过朋友投了星海科技,数据分析师\n",
        encoding="utf-8",
    )
    parsed = parse_standard_workbook(source)
    calls = []

    class Service:
        def parse_standard_positions(self, *_args, **_kwargs):
            raise AssertionError("非标准表格不得走确定性标准表入口")

        async def parse_batch(self, user_id, text):
            calls.append((user_id, text))
            return {
                "status": "preview",
                "positions": [{
                    "company": "星海科技",
                    "position": "数据分析师",
                    "stage": "applied",
                    "department": None,
                    "channel": None,
                    "next_action": None,
                    "already_exists": False,
                }],
            }

    response = asyncio.run(ParseJobsTool(Service(), "u1").arun({
        "text": parsed.structured_text,
    }))

    assert parsed.is_standard is False
    assert standard_positions_from_structured_text(parsed.structured_text) is None
    assert response.status == "success"
    assert len(calls) == 1
    assert calls[0][0] == "u1"
    assert "A2「我的记录」=上周通过朋友投了星海科技" in calls[0][1]
    assert "B2「目标」=数据分析师" in calls[0][1]
